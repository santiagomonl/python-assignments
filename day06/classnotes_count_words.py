##Count_words
def count_words(text):
    words = text.split(",") #split the string where...
    return len(words)

print(count_words("name,23,value,42.1"))

def count_number_of_each_words(text):
    words = text.split(",")
    word_count = {}
    for word in words:
        if word in word_count:
            word_count[word] += 1
        else:
            word_count[word] = 1
    return word_count

res = count_number_of_each_words("name,23,value,42.1,name,25")
# print(res)
# print(res["name"])

# for key in res.keys():
#     print (key, res[key])

# for key in ["23", "42.1", "25", "100"]:
#     if key in res:
#         print(key, res[key])
#     else:
#         print(key, "Not found")
#     print(key, res.get(key, "Not found")) #This does the same of just taking the values that are in the dic otherwise fives "not found" message

# def count_number_of_each_words_using_default_dict(text):
#     from collections import defaultdict
#     words = text.split(",")
#     word_count = defaultdict(int)
#     for word in words:
#         word_count[word] += 1
#     return word_count

#Tuple are list that cannot be change, you can have them from x.items
# for pair in res.items():
#     print(pair)
#     print(pair[0], pair[1])

# #but one can also call individually the two paired string and value
# for key, value in res.items():
#     print(key, value)

# #Overlapping elements in a set

# #Opening files
# with open(filename, "r") as fh:
#     fh.read()
#     fh.readline() #read from a simple text file

# #Image reading
# from PIL import Image
# image= Image.open()

#read an excel file
import openpyxl #might require installation
import sys

if len(sys.argv) !=2:
    exit(f"Usage: {sys.argv[0]} FILENAME")

in_file = sys.argv[1]

wb = openpyxl.load.workbook(filename = in_file)
for ws in wb.worksheets:
    print(ws.title)

ws = wb.worksheets[0] #Uses the first tab (worksheet), i can add input to just call in numbner which tab
print(ws["A1"].value)

def read_excelwith_openpyxl(file_path):
    from openpyxl import load_workbook
    wb= load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value, end=' ')
        print()
    wb.close()

def read_excel(file_path):
    import pandas as pd
    df= pd.read_excel(file_path)
    print(df)
    df.to_excel("output.xlsx", index=False)
    df.to
